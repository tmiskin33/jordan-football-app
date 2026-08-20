"""
Generates the two blank charting workbooks the site accepts.

Headers here are the exact strings src/lib/excelImport.ts looks up, and the
sheet names are the exact ones it opens. Columns are matched by header NAME,
so they can be reordered or have extras added, but a renamed header stops
being read.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

MAROON = "730F37"
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", fgColor=MAROON)
TITLE_FONT = Font(bold=True, size=14, color=MAROON)
NOTE_FONT = Font(size=10)
BOLD = Font(bold=True, size=10)

# Accepted values, from the workbooks' own Lists sheets.
VALUES = {
    "Hash": ["L", "M", "R"],
    "Direction": ["L", "M", "R"],
    "Play Type": ["Run", "Pass"],
    "Strength": ["L", "R", "BAL"],
    "Score Situation": ["Ahead", "Tied", "Behind"],
    "Backfield": ["Under Center", "Shotgun", "Pistol", "Empty"],
    "Motion (Y/N)": ["Y", "N"],
    "Blitz (Y/N)": ["Y", "N"],
    "Personnel": ["10", "11", "12", "13", "20", "21", "22", "23"],
    "Front": ["4-3", "3-4", "Nickel", "Dime", "5-2", "3-3 Stack", "Bear", "Goal Line"],
    "Coverage": ["Cover 0", "Cover 1", "Cover 2", "Cover 2 Man", "Cover 3", "Cover 4",
                 "Cover 6", "Man Free"],
    "Blitz Type": ["Edge Pressure", "A-Gap", "Corner Blitz", "Safety Blitz", "Zone Blitz",
                   "Fire Zone"],
    "Result Type": ["Rush", "Complete", "Incomplete", "TD", "INT", "Sack", "Sack, Fumble",
                    "Fumble", "Fumble Lost", "Scramble", "Penalty", "First Down", "Gain",
                    "No Gain", "Loss", "Timeout"],
    "ODK": ["K", "S"],
}

# Which dropdown list (if any) each header uses.
DROPDOWN_FOR = {
    "Hash": "Hash", "Direction": "Direction", "Off Direction": "Direction",
    "Play Type": "Play Type", "Offense Play Type": "Play Type",
    "Strength": "Strength", "Off Strength": "Strength",
    "Score Situation": "Score Situation", "Backfield": "Backfield",
    "Motion (Y/N)": "Motion (Y/N)", "Blitz (Y/N)": "Blitz (Y/N)",
    "Personnel": "Personnel", "Def Personnel": "Personnel",
    "Front": "Front", "Def Front": "Front", "Our Front": "Front",
    "Coverage": "Coverage", "Our Coverage": "Coverage",
    "Blitz Type": "Blitz Type", "Result Type": "Result Type", "ODK": "ODK",
}

OPP_OFFENSE = ["Film / Game", "Qtr", "Down", "Distance", "Yard Line (to goal)", "Hash",
               "Score Situation", "Personnel", "Formation", "Strength", "Backfield",
               "Motion (Y/N)", "Play Type", "Play Call / Concept", "Direction",
               "Ball Carrier / Target", "Yards", "Result Type", "Front", "Coverage",
               "Blitz (Y/N)", "Blitz Type"]

OPP_DEFENSE = ["Film / Game", "Qtr", "Down", "Distance", "Yard Line (to goal)", "Hash",
               "Def Personnel", "Front", "Coverage", "Blitz (Y/N)", "Blitz Type",
               "Rushers (#)", "Offense Play Type", "Formation Faced", "Off Strength",
               "Off Direction", "Yards Allowed", "Result Type"]

OPP_ST = ["Film / Game", "Play #", "ODK", "Qtr", "Down", "Distance",
          "Yard Line (to goal)", "Hash", "Play Type", "Result", "Yards"]

TEAM_OFFENSE = ["Week", "Date", "Opponent", "Qtr", "Down", "Distance",
                "Yard Line (to goal 1-99)", "Hash", "Personnel", "Formation", "Strength",
                "Play Type", "Play Call / Concept", "Direction", "Key Player", "Yards",
                "Result Type", "Points", "Def Front", "Coverage", "Blitz (Y/N)"]

TEAM_DEFENSE = ["Week", "Date", "Opponent", "Qtr", "Down", "Distance",
                "Yard Line (to goal 1-99)", "Hash", "Personnel", "Opp Formation",
                "Strength", "Play Type", "Opp Play Call", "Direction", "Opp Key Player",
                "Yards Allowed", "Result Type", "Points", "Our Front", "Our Coverage",
                "Blitz (Y/N)"]

TEAM_ST = ["Week", "Opponent", "Play #", "ODK", "Qtr", "Down", "Distance",
           "Yard Line (to goal)", "Hash", "Play Type", "Result", "Yards"]


def write_log_sheet(wb, name, headers, required_note):
    ws = wb.create_sheet(name)
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = max(12, min(len(h) + 4, 26))
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Dropdowns down a generous range so charting can just keep going.
    for c, h in enumerate(headers, start=1):
        key = DROPDOWN_FOR.get(h)
        if not key:
            continue
        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(VALUES[key]) + '"',
            allow_blank=True,
            showDropDown=False,  # False = DO show the in-cell dropdown arrow
        )
        ws.add_data_validation(dv)
        col = get_column_letter(c)
        dv.add(f"{col}2:{col}1000")

    ws.cell(row=1, column=len(headers) + 2, value=required_note).font = NOTE_FONT
    return ws


def write_readme(wb, title, lines):
    ws = wb.create_sheet("Read Me", 0)
    ws.column_dimensions["B"].width = 4
    ws.column_dimensions["C"].width = 110
    ws.cell(row=2, column=3, value=title).font = TITLE_FONT
    r = 4
    for kind, text in lines:
        cell = ws.cell(row=r, column=3, value=text)
        if kind == "h":
            cell.font = BOLD
            r += 1
        elif kind == "b":
            cell.font = NOTE_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = max(15, 15 * (len(text) // 105 + 1))
            r += 1
        else:
            r += 1
    return ws


def write_values_sheet(wb):
    ws = wb.create_sheet("Valid Values")
    ws.cell(row=1, column=2, value="Accepted values").font = TITLE_FONT
    ws.cell(row=2, column=2,
            value="Anything else still imports as free text — these are just the "
                  "vocabulary the dashboards group by.").font = NOTE_FONT
    col = 2
    for key, vals in VALUES.items():
        ws.cell(row=4, column=col, value=key).font = HEADER_FONT
        ws.cell(row=4, column=col).fill = HEADER_FILL
        for i, v in enumerate(vals, start=5):
            ws.cell(row=i, column=col, value=v).font = NOTE_FONT
        ws.column_dimensions[get_column_letter(col)].width = 18
        col += 1
    return ws


# ---------------------------------------------------------------- opponent
wb = openpyxl.Workbook()
wb.remove(wb.active)

write_log_sheet(wb, "Opp Offense Log", OPP_OFFENSE,
                "REQUIRED on every row: Film / Game.")
write_log_sheet(wb, "Opp Defense Log", OPP_DEFENSE,
                "REQUIRED on every row: Film / Game.")
write_log_sheet(wb, "Opp Special Teams Log", OPP_ST,
                "REQUIRED on every row: Film / Game AND Play #.")
write_values_sheet(wb)
write_readme(wb, "Opponent Scouting Template", [
    ("b", "Chart an opponent's film here, then upload this file on that opponent's page. "
          "One workbook can hold several games of their film."),
    ("", ""),
    ("h", "HOW TO USE IT"),
    ("b", "1. Chart one row per snap on the three log sheets. Only the sheets you use need rows."),
    ("b", "2. On the site, go to Scouting, pick the opponent, and click 'Choose workbook to import'."),
    ("b", "3. Every tendency, concept and explosive-play table rebuilds itself from these rows."),
    ("", ""),
    ("h", "THE ONE RULE THAT MATTERS"),
    ("b", "Film / Game must be filled on every row. It is how snaps get grouped into separate "
          "films, and rows without it are skipped. Use whatever names you like, but keep them "
          "identical within a game — 'vs Juab' and 'Vs Juab' become two different films."),
    ("", ""),
    ("h", "COLUMNS"),
    ("b", "Columns are matched by the header text in row 1, not by position. You can reorder them "
          "or add your own extra columns, but do not rename these headers or they stop being read."),
    ("b", "Every column except Film / Game is optional. Anything you leave blank simply reads as "
          "'not charted' — the tables that need it will say so rather than guessing."),
    ("", ""),
    ("h", "HOW TO FILL THE TRICKY ONES"),
    ("b", "Yard Line (to goal) — 1 to 99, counting down to the goal the OFFENSE is attacking. "
          "Their own 20 is 80; midfield is 50; first-and-goal from the 5 is 5."),
    ("b", "Play Type — Run or Pass. A snap with no Play Type stays in the log but is left out of "
          "the tendency percentages, which is how timeouts and dead-ball penalties are handled."),
    ("b", "Result Type — use TD for any score and INT for any interception, otherwise the "
          "touchdown and turnover counts miss them. Use 'Fumble Lost' only when the film shows "
          "the ball was actually recovered by the defense."),
    ("b", "Yards — the gain or loss on the play. Negative numbers are fine. An interception with "
          "no yardage still counts as a failed play."),
    ("", ""),
    ("h", "WORKED EXAMPLE — one row on Opp Offense Log"),
    ("b", "Film / Game: vs Juab   |   Qtr: 1   |   Down: 2   |   Distance: 7   |   "
          "Yard Line (to goal): 62   |   Hash: L"),
    ("b", "Formation: TRIO   |   Strength: R   |   Backfield: Shotgun   |   Motion (Y/N): N   |   "
          "Personnel: 11"),
    ("b", "Play Type: Pass   |   Play Call / Concept: MESH   |   Direction: R   |   "
          "Yards: 14   |   Result Type: Complete"),
    ("b", "Reads as: 2nd & Long, own side of the field, out of TRIO — a 14-yard completion. "
          "Counts as a success (7 to go on 2nd down needs 4.9) but not explosive (a pass needs 15)."),
    ("", ""),
    ("h", "WHAT THE SITE WORKS OUT ON ITS OWN"),
    ("b", "Do not add columns for these — they are recalculated on every import, so a stale value "
          "in the sheet would just be ignored: success, explosive, down & distance bucket, field "
          "zone, run/pass rate, tendency, yards per play, conversion rate, stop rate."),
    ("b", "Success = gaining 50% of the distance on 1st down, 70% on 2nd, all of it on 3rd/4th. "
          "Explosive = a run of 10+ or a pass of 15+."),
])
wb.save(r"C:\Users\tanne\Football Project\jordan-football-app\templates\Opponent Scouting Template.xlsx")

# ------------------------------------------------------------------- team
wb2 = openpyxl.Workbook()
wb2.remove(wb2.active)

write_log_sheet(wb2, "Offense Play-by-Play", TEAM_OFFENSE,
                "REQUIRED on every row: Down or Play Type.")
write_log_sheet(wb2, "Defense Play-by-Play", TEAM_DEFENSE,
                "REQUIRED on every row: Down or Play Type.")
write_log_sheet(wb2, "Special Teams Log", TEAM_ST,
                "REQUIRED on every row: Play #.")
write_values_sheet(wb2)
write_readme(wb2, "Team Analytics Template  (our own team, one game per file)", [
    ("b", "Chart Jordan's own snaps here — one file per game. Upload it from that game's "
          "chart/film page and it feeds Team Analytics and the season dashboard."),
    ("", ""),
    ("h", "HOW TO USE IT"),
    ("b", "1. Chart one row per snap: our offensive snaps on the first sheet, the snaps our "
          "defense faced on the second."),
    ("b", "2. On the site, open Team Analytics, find the game, and click 'chart/film'."),
    ("b", "3. Click 'Choose workbook to import' and pick this file."),
    ("", ""),
    ("h", "ONE FILE PER GAME"),
    ("b", "This format is always about a single game, so the site ties everything in it to the "
          "game you upload it from. That is also why it has to be uploaded from a game page and "
          "not from the general import — it would have no way to know which week it belongs to."),
    ("", ""),
    ("h", "WHOSE SIDE IS WHICH"),
    ("b", "On Offense Play-by-Play everything is ours: our formation, our play call, our key "
          "player. Def Front / Coverage / Blitz describe what the DEFENSE showed us."),
    ("b", "On Defense Play-by-Play it flips: Opp Formation and Opp Play Call are theirs, while "
          "Our Front, Our Coverage and Blitz are ours. Yards Allowed is what they gained."),
    ("", ""),
    ("h", "COLUMNS"),
    ("b", "Columns are matched by the header text in row 1, not by position. Reorder or add your "
          "own freely; just do not rename these headers."),
    ("b", "Week / Date / Opponent are yours to keep the file self-describing — the site takes the "
          "game from where you upload it, so they can stay blank."),
    ("b", "A row needs at least a Down or a Play Type to count as a snap. Everything else is "
          "optional and reads as 'not charted' when blank."),
    ("", ""),
    ("h", "HOW TO FILL THE TRICKY ONES"),
    ("b", "Yard Line (to goal 1-99) — counting down to the goal the offense on that snap is "
          "attacking. On our offensive sheet that is the goal WE are attacking; on the defensive "
          "sheet it is the one THEY are attacking."),
    ("b", "Result Type — TD for any score, INT for any interception. On the defensive sheet an "
          "INT is a takeaway for us."),
    ("b", "Key Player — the ball carrier or target. Fill it and the player-usage table comes to "
          "life; leave it blank and that table stays empty."),
    ("", ""),
    ("h", "WORKED EXAMPLE — one row on Offense Play-by-Play"),
    ("b", "Qtr: 3   |   Down: 3   |   Distance: 4   |   Yard Line (to goal 1-99): 38   |   Hash: M"),
    ("b", "Formation: DOUBLES   |   Strength: L   |   Personnel: 11   |   Play Type: Run   |   "
          "Play Call / Concept: 34 GATOR   |   Direction: L"),
    ("b", "Key Player: T. Smith   |   Yards: 12   |   Result Type: First Down"),
    ("b", "Reads as: 3rd & Medium in the fringe — a 12-yard run. Converts the third down, counts "
          "as a success, and is explosive (a run needs 10)."),
    ("", ""),
    ("h", "WHAT THE SITE WORKS OUT ON ITS OWN"),
    ("b", "Do not add columns for these — they are recalculated on every import: success, "
          "explosive, down & distance bucket, field zone, run/pass rate, yards per play, third "
          "down conversion rate, stop rate, takeaways, and every season-dashboard figure."),
    ("b", "Success = gaining 50% of the distance on 1st down, 70% on 2nd, all of it on 3rd/4th. "
          "Explosive = a run of 10+ or a pass of 15+."),
])
wb2.save(r"C:\Users\tanne\Football Project\jordan-football-app\templates\Team Analytics Template.xlsx")

print("wrote both templates")
